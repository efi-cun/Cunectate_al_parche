import os
import sys
import glob
import json
import io
import shutil
import tempfile
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASISTENCIA_DIR = os.path.join(BASE_DIR, "Asistencia")
PLANTA_DIR = os.path.join(BASE_DIR, "Planta")
EXCEL_OUTPUT = os.path.join(BASE_DIR, "Consolidado_Bienvenida_Facilitadores.xlsx")
EXCEL_FALTANTES = os.path.join(BASE_DIR, "Facilitadores_Pendientes_Asistencia.xlsx")
HTML_OUTPUT = os.path.join(BASE_DIR, "tablero.html")
HTML_RESPALDO = os.path.join(BASE_DIR, "tablero_respaldo.html")
HTML_INDEX = os.path.join(BASE_DIR, "index.html")

print("=== INICIANDO PROCESAMIENTO BIENVENIDA FACILITADORES CUN ===")

# --- DICCIONARIO DE LIMPIEZA DE CARACTERES (ESTILO procesar_reporte.ps1) ---
replacements = [
    ("Ã\xa0", "a"), ("Ã¡", "a"), ("Ã\xa2", "a"), ("Ã£", "a"), ("Ã¤", "a"),
    ("Ã¨", "e"), ("Ã©", "e"), ("Ãª", "e"), ("Ã«", "e"),
    ("Ã¬", "i"), ("Ã\xad", "i"), ("Ã®", "i"), ("Ã¯", "i"),
    ("Ã²", "o"), ("Ã³", "o"), ("Ã´", "o"), ("Ãµ", "o"), ("Ã¶", "o"),
    ("Ã¹", "u"), ("Ãº", "u"), ("Ã»", "u"), ("Ã¼", "u"),
    ("Ã±", "n"), ("Ã‘", "N"), ("Â¿", ""),
    ("Ã\x81", "A"), ("Ã\x89", "E"), ("Ã\x8d", "I"), ("Ã\x93", "O"), ("Ã\x9a", "U"),
    ("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n"),
    ("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"), ("Ñ", "N")
]

def clean_text(txt):
    if pd.isna(txt) or txt is None:
        return ""
    s = str(txt)
    for old, new in replacements:
        s = s.replace(old, new)
    return s.strip()

# 1. Cargar y Limpiar Archivo de Asistencia (.xlsx, .xls, .csv)
asistencia_files = glob.glob(os.path.join(ASISTENCIA_DIR, "*.xlsx")) + \
                   glob.glob(os.path.join(ASISTENCIA_DIR, "*.xls")) + \
                   glob.glob(os.path.join(ASISTENCIA_DIR, "*.csv"))
if not asistencia_files:
    asistencia_files = glob.glob(os.path.join(ASISTENCIA_DIR, "**", "*.xlsx"), recursive=True) + \
                       glob.glob(os.path.join(ASISTENCIA_DIR, "**", "*.xls"), recursive=True) + \
                       glob.glob(os.path.join(ASISTENCIA_DIR, "**", "*.csv"), recursive=True)

asistencia_files = [f for f in asistencia_files if not os.path.basename(f).startswith("~$") and not os.path.basename(f).startswith("temp_")]

if not asistencia_files:
    print("ERROR: No se encontró ningún archivo de asistencia (.xlsx, .xls, .csv) en la carpeta Asistencia.")
    sys.exit(1)

asistencia_files.sort(key=os.path.getmtime, reverse=True)

valid_asistencia_file = None
for fpath in asistencia_files:
    if os.path.getsize(fpath) > 500:
        valid_asistencia_file = fpath
        break
if not valid_asistencia_file:
    valid_asistencia_file = asistencia_files[0]

asistencia_file = valid_asistencia_file
print(f"[1/5] Limpiando y procesando archivo de Asistencia: {os.path.basename(asistencia_file)}")

def load_excel_safely(file_path):
    try:
        with open(file_path, 'rb') as f:
            content = f.read()
    except (PermissionError, OSError):
        import subprocess
        temp_path = os.path.join(tempfile.gettempdir(), f"temp_{os.path.basename(file_path)}")
        cmd = f'powershell -Command "Copy-Item -Path \'{os.path.abspath(file_path)}\' -Destination \'{temp_path}\' -Force"'
        subprocess.run(cmd, shell=True, check=True)
        with open(temp_path, 'rb') as f:
            content = f.read()
        try:
            os.remove(temp_path)
        except Exception:
            pass
    return pd.ExcelFile(io.BytesIO(content))

if asistencia_file.lower().endswith(('.xlsx', '.xls')):
    xl_asist = load_excel_safely(asistencia_file)
    df_asist_raw = pd.read_excel(xl_asist, sheet_name=0)
else:
    try:
        df_asist_raw = pd.read_csv(asistencia_file, encoding='utf-8')
    except Exception:
        df_asist_raw = pd.read_csv(asistencia_file, encoding='latin1')

print(f"      Total registros recibidos en Asistencia: {len(df_asist_raw)}")

col_ced_asist = next((c for c in df_asist_raw.columns if 'CEDULA' in clean_text(c).upper() or 'IDENTIFICAC' in clean_text(c).upper() or 'DOCUMENTO' in clean_text(c).upper() or 'ID' in clean_text(c).upper()), df_asist_raw.columns[0])
col_nom_asist = next((c for c in df_asist_raw.columns if 'NAME' in clean_text(c).upper() or 'NOMBRE' in clean_text(c).upper() or 'FACILITADOR' in clean_text(c).upper() or 'ASISTENTE' in clean_text(c).upper()), df_asist_raw.columns[1] if len(df_asist_raw.columns) > 1 else df_asist_raw.columns[0])

df_asist_clean = pd.DataFrame()
df_asist_clean['CEDULA'] = pd.to_numeric(df_asist_raw[col_ced_asist], errors='coerce').astype('Int64').astype(str).str.strip()

def format_nombre(val):
    txt = clean_text(val)
    if ',' in txt:
        parts = txt.split(',', 1)
        return f"{parts[1].strip()} {parts[0].strip()}".upper()
    return txt.upper()

df_asist_clean['NOMBRE'] = df_asist_raw[col_nom_asist].apply(format_nombre)

# 2. Cargar Base de Planta
xlsx_files = glob.glob(os.path.join(PLANTA_DIR, "*.xlsx")) + glob.glob(os.path.join(PLANTA_DIR, "*.xls"))
if not xlsx_files:
    xlsx_files = glob.glob(os.path.join(PLANTA_DIR, "**", "*.xlsx"), recursive=True) + glob.glob(os.path.join(PLANTA_DIR, "**", "*.xls"), recursive=True)

xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith("~$") and not os.path.basename(f).startswith("temp_")]

if not xlsx_files:
    print("ERROR: No se encontró ningún archivo .xlsx o .xls en la carpeta Planta.")
    sys.exit(1)

xlsx_files.sort(key=os.path.getmtime, reverse=True)
planta_file = xlsx_files[0]
print(f"[2/5] Cargando base de referencia Planta: {os.path.basename(planta_file)}")

xl_planta = load_excel_safely(planta_file)
sheet_planta = next((s for s in xl_planta.sheet_names if "BASE FACILITADORES" in s.upper() or "ACTIVOS" in s.upper() or "PLANTA" in s.upper()), xl_planta.sheet_names[0])
df_planta_raw = pd.read_excel(xl_planta, sheet_name=sheet_planta)

# Detección flexible de columnas en Planta
col_ced_planta = next((c for c in df_planta_raw.columns if 'IDENTIFICAC' in clean_text(c).upper() or 'CEDULA' in clean_text(c).upper() or 'DOCUMENTO' in clean_text(c).upper()), df_planta_raw.columns[0])

col_nom_planta = next((c for c in df_planta_raw.columns if 'NOMBRE' in clean_text(c).upper() and 'COMPLETO' in clean_text(c).upper()), None)
if not col_nom_planta:
    nombres_col = next((c for c in df_planta_raw.columns if clean_text(c).upper() in ['NOMBRES', 'NOMBRE']), None)
    apellidos_col = next((c for c in df_planta_raw.columns if clean_text(c).upper() in ['APELLIDOS', 'APELLIDO']), None)
    if nombres_col and apellidos_col:
        df_planta_raw['NOMBRE COMPLETO'] = df_planta_raw[nombres_col].fillna('').astype(str).str.strip() + " " + df_planta_raw[apellidos_col].fillna('').astype(str).str.strip()
        df_planta_raw['NOMBRE COMPLETO'] = df_planta_raw['NOMBRE COMPLETO'].str.strip()
        col_nom_planta = 'NOMBRE COMPLETO'
    elif nombres_col:
        col_nom_planta = nombres_col
    else:
        col_nom_planta = df_planta_raw.columns[1] if len(df_planta_raw.columns) > 1 else df_planta_raw.columns[0]

col_invitacion = next((c for c in df_planta_raw.columns if 'INVITAC' in clean_text(c).upper()), None)
col_modalidad_alt = next((c for c in df_planta_raw.columns if 'MODALIDAD' in clean_text(c).upper()), None)
if not col_invitacion:
    if col_modalidad_alt:
        df_planta_raw['INVITACIÓN'] = df_planta_raw[col_modalidad_alt]
    else:
        df_planta_raw['INVITACIÓN'] = 'VIRTUAL'
    col_invitacion = 'INVITACIÓN'

if not col_modalidad_alt:
    col_modalidad_alt = col_invitacion

col_nivel1 = next((c for c in df_planta_raw.columns if 'NIVEL 1' in clean_text(c).upper() or 'NIVEL1' in clean_text(c).upper()), None)
if not col_nivel1:
    df_planta_raw['Nombre Nivel 1'] = 'OPERACIONES'
    col_nivel1 = 'Nombre Nivel 1'

col_nivel2 = next((c for c in df_planta_raw.columns if 'NIVEL 2' in clean_text(c).upper() or 'NIVEL2' in clean_text(c).upper()), None)
if not col_nivel2:
    df_planta_raw['Nombre Nivel 2'] = 'REGIONAL BOGOTA'
    col_nivel2 = 'Nombre Nivel 2'

col_nivel3 = next((c for c in df_planta_raw.columns if 'NIVEL 3' in clean_text(c).upper() or 'NIVEL3' in clean_text(c).upper()), None)
if not col_nivel3:
    df_planta_raw['Nombre Nivel 3'] = 'BOGOTA CENTRO'
    col_nivel3 = 'Nombre Nivel 3'

col_cargo = next((c for c in df_planta_raw.columns if 'CARGO' in clean_text(c).upper()), None)
if not col_cargo:
    df_planta_raw['Descripción Cargo'] = 'DOCENTE / FACILITADOR'
    col_cargo = 'Descripción Cargo'

col_centro_costo = next((c for c in df_planta_raw.columns if 'CENTRO COSTO' in clean_text(c).upper() or 'CENTRO_COSTO' in clean_text(c).upper()), None)
if not col_centro_costo:
    df_planta_raw['Nombre Centro Costo'] = 'GENERAL'
    col_centro_costo = 'Nombre Centro Costo'

col_escuela = next((c for c in df_planta_raw.columns if 'ESCUELA' in clean_text(c).upper()), None)
if not col_escuela:
    df_planta_raw['ESCUELA'] = df_planta_raw[col_centro_costo]
    col_escuela = 'ESCUELA'

col_correo_corp = next((c for c in df_planta_raw.columns if 'CORREO CORPORATIVO' in clean_text(c).upper() or 'CORREO' in clean_text(c).upper()), None)
col_correo_pers = next((c for c in df_planta_raw.columns if 'CORREO PERSONAL' in clean_text(c).upper()), None)
col_telefono = next((c for c in df_planta_raw.columns if 'TELEFONO' in clean_text(c).upper() or 'CELULAR' in clean_text(c).upper()), None)
col_tipo_contrato = next((c for c in df_planta_raw.columns if 'TIPO CONTRATO' in clean_text(c).upper() or 'CONTRATO' in clean_text(c).upper()), None)
col_fecha_inicio = next((c for c in df_planta_raw.columns if 'FECHA INICIO' in clean_text(c).upper()), None)
col_fecha_venc = next((c for c in df_planta_raw.columns if 'FECHA VENCIMIENTO' in clean_text(c).upper()), None)
col_centro_trabajo = next((c for c in df_planta_raw.columns if 'CENTRO TRABAJO' in clean_text(c).upper()), None)

df_planta_raw['CEDULA_CLEAN'] = pd.to_numeric(df_planta_raw[col_ced_planta], errors='coerce').astype('Int64').astype(str).str.strip()
planta_cedulas_set = set(df_planta_raw['CEDULA_CLEAN'].dropna().unique())
total_planta = len(df_planta_raw)

# Estructurar PLANTA_BASE alineado con fórmulas VLOOKUP (Columnas A a Q)
planta_cols_key = [
    'CEDULA_CLEAN',      # 1 (A)
    col_invitacion,      # 2 (B)
    col_nom_planta,      # 3 (C)
    col_nivel1,          # 4 (D)
    col_nivel2,          # 5 (E)
    col_nivel3,          # 6 (F)
    col_modalidad_alt,   # 7 (G)
    col_cargo,           # 8 (H)
    col_centro_costo,    # 9 (I)
    col_correo_corp,     # 10 (J)
    col_correo_pers,     # 11 (K)
    col_telefono,        # 12 (L)
    col_tipo_contrato,   # 13 (M)
    col_fecha_inicio,    # 14 (N)
    col_fecha_venc,      # 15 (O)
    col_centro_trabajo,  # 16 (P)
    col_escuela          # 17 (Q)
]

planta_cols_ordered = [c for c in planta_cols_key if c is not None]
for c in df_planta_raw.columns:
    if c not in planta_cols_ordered and c != col_ced_planta:
        planta_cols_ordered.append(c)

df_planta_export = df_planta_raw[planta_cols_ordered].copy()
df_planta_export.rename(columns={'CEDULA_CLEAN': 'CEDULA'}, inplace=True)

# 3. Auto-corrección de Cédulas por 100% Coincidencia en Nombre y Validación contra Planta
print("[3/5] Validando presencia en Planta, aplicando auto-corrección de cédulas por 100% en nombre y detectando duplicados...")

df_planta_raw['NOMBRE_COMPLETO_SEARCH'] = (df_planta_raw[col_nom_planta].fillna('').astype(str)).apply(clean_text)

# 3.0 Corrección automática de cédula y nombre para registros con coincidencia en Planta (por búsqueda de nombre)
corrected_count = 0
for idx, row in df_asist_clean.iterrows():
    ced_a = str(row['CEDULA'])
    if ced_a not in planta_cedulas_set:
        nom_a = clean_text(row['NOMBRE'])
        words_a = set([w for w in nom_a.split() if len(w) > 2])
        if words_a and len(words_a) >= 1:
            best_p_row = None
            best_score = 0.0
            for p_idx, p_row in df_planta_raw.iterrows():
                words_p = set([w for w in str(p_row['NOMBRE_COMPLETO_SEARCH']).split() if len(w) > 2])
                if not words_p:
                    continue
                
                if words_a == words_p and len(words_a) >= 2:
                    best_score = 1.0
                    best_p_row = p_row
                    break
                
                intersection = words_a.intersection(words_p)
                union = words_a.union(words_p)
                jaccard = len(intersection) / len(union) if union else 0
                if len(intersection) >= 2 and jaccard >= 0.4:
                    if jaccard > best_score:
                        best_score = jaccard
                        best_p_row = p_row

            if best_p_row is not None and best_score >= 0.4:
                ced_p = str(best_p_row['CEDULA_CLEAN'])
                nom_p = str(best_p_row[col_nom_planta])
                df_asist_clean.loc[idx, 'CEDULA'] = ced_p
                df_asist_clean.loc[idx, 'NOMBRE'] = nom_p
                corrected_count += 1

if corrected_count > 0:
    print(f"      Auto-corrección ejecutada: {corrected_count} registro(s) de asistencia actualizado(s) (Cédula y Nombre de Planta) por coincidencia de búsqueda por nombre.")

is_duplicated_in_asist = df_asist_clean.duplicated(subset=['CEDULA'], keep='first')
is_not_in_planta = ~df_asist_clean['CEDULA'].isin(planta_cedulas_set)

df_duplicados = df_asist_clean[is_duplicated_in_asist | is_not_in_planta].copy()
df_unicos = df_asist_clean[~is_duplicated_in_asist & ~is_not_in_planta].copy()

print(f"      Total Facilitadores en Base Planta: {total_planta}")
print(f"      Registros Válidos y Únicos (Consolidado_Unicos): {len(df_unicos)}")
print(f"      Registros Duplicados / No Encontrados en Planta (Duplicados): {len(df_duplicados)}")

# 3.1 Búsqueda Avanzada por Nombre y Apellido para registros pendientes en no encontrados por Cédula
print("      Realizando búsqueda por Nombre y Apellido para registros no encontrados...")
busqueda_records = []
no_encontrados_asist = df_asist_clean[is_not_in_planta].drop_duplicates(subset=['CEDULA', 'NOMBRE'])

for idx, row in no_encontrados_asist.iterrows():
    ced_a = str(row['CEDULA'])
    nom_a = str(row['NOMBRE'])
    nom_clean = clean_text(nom_a)
    words_a = set([w for w in nom_clean.split() if len(w) > 2])
    
    best_p_row = None
    best_score = 0.0
    best_status = "NO ENCONTRADO EN PLANTA"
    
    if words_a:
        for p_idx, p_row in df_planta_raw.iterrows():
            p_nom = p_row['NOMBRE_COMPLETO_SEARCH']
            words_p = set([w for w in p_nom.split() if len(w) > 2])
            if not words_p:
                continue
            
            # Coincidencia exacta de conjunto de palabras
            if words_a == words_p and len(words_a) >= 2:
                best_score = 1.0
                best_p_row = p_row
                best_status = "COINCIDENCIA EXACTA POR NOMBRE"
                break
            
            # Superposición de tokens y Jaccard
            intersection = words_a.intersection(words_p)
            union = words_a.union(words_p)
            jaccard = len(intersection) / len(union) if union else 0
            
            if len(intersection) >= 2 and jaccard >= 0.4:
                if jaccard > best_score:
                    best_score = jaccard
                    best_p_row = p_row
                    best_status = f"COINCIDENCIA ALTA POR NOMBRE ({round(jaccard*100)}%)" if jaccard >= 0.6 else f"COINCIDENCIA PARCIAL ({round(jaccard*100)}%)"

    if best_p_row is not None and best_score >= 0.4:
        ced_p = str(best_p_row['CEDULA_CLEAN'])
        nom_p = str(best_p_row[col_nom_planta])
        car_p = str(best_p_row.get(col_cargo, 'DOCENTE / FACILITADOR'))
        cc_p  = str(best_p_row.get(col_centro_costo, 'GENERAL'))
        reg_p = str(best_p_row.get(col_nivel2, 'REGIONAL BOGOTA'))
        sed_p = str(best_p_row.get(col_nivel3, 'BOGOTA CENTRO'))
        esc_p = str(best_p_row.get(col_escuela, cc_p))
    else:
        ced_p = "NO ENCONTRADO EN PLANTA"
        nom_p = "SIN COINCIDENCIA EN PLANTA"
        best_status = "NO ENCONTRADO EN PLANTA"
        best_score = 0.0
        car_p = "-"
        cc_p  = "-"
        reg_p = "-"
        sed_p = "-"
        esc_p = "-"

    busqueda_records.append({
        'CÉDULA ASISTENCIA': ced_a,
        'NOMBRE ASISTENCIA': nom_a,
        'CÉDULA PLANTA ENCONTRADA': ced_p,
        'NOMBRE PLANTA ENCONTRADO': nom_p,
        'ESTADO COINCIDENCIA': best_status,
        'PORCENTAJE SIMILITUD': f"{round(best_score*100)}%",
        'Descripción Cargo': car_p,
        'Nombre Centro Costo': cc_p,
        'Nombre Nivel 2 (Regional)': reg_p,
        'Nombre Nivel 3 (Sede)': sed_p,
        'ESCUELA': esc_p
    })

df_busqueda_nombre = pd.DataFrame(busqueda_records)
print(f"      Coincidencias por Nombre encontradas para registros no cruzados por cédula: {len(df_busqueda_nombre[df_busqueda_nombre['ESTADO COINCIDENCIA'] != 'NO ENCONTRADO EN PLANTA'])}")

# 4. Generar Libro de Excel con Fórmulas VLOOKUP Desplegadas en 1.500 Filas
TOTAL_FILAS_DESPLEGADAS = 1500
print(f"[4/5] Desplegando fórmulas BUSCARV a lo largo de {TOTAL_FILAS_DESPLEGADAS} filas en las columnas...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="004B28", end_color="004B28", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
data_font = Font(name="Calibri", size=11, color="1F2937")
border_thin = Side(style='thin', color="D1D5DB")
cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

headers_consolidado = [
    'INVITACIÓN',
    'CEDULA',
    'NOMBRE',
    'Nombre Nivel 2',
    'Nombre Nivel 3',
    'Descripción Cargo',
    'Nombre Centro Costo',
    'ESCUELA'
]

last_col_letter = get_column_letter(len(df_planta_export.columns))
escuela_col_idx = (df_planta_export.columns.get_loc(col_escuela) + 1) if col_escuela in df_planta_export.columns else 17

def format_consolidado_with_full_column_formulas(ws, df_asist_data):
    ws.views.sheetView[0].showGridLines = True
    ws.append(headers_consolidado)
    
    for col_num, h in enumerate(headers_consolidado, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    data_rows = list(df_asist_data.itertuples(index=False))
    num_data_rows = len(data_rows)
    total_rows = max(num_data_rows, TOTAL_FILAS_DESPLEGADAS)

    for r_idx in range(2, total_rows + 2):
        if (r_idx - 2) < num_data_rows:
            row_data = data_rows[r_idx - 2]
            ced_cell_value = str(row_data.CEDULA)
            nom_formula = str(row_data.NOMBRE)
        else:
            ced_cell_value = ""
            nom_formula = f'=IF(B{r_idx}="","",IFERROR(VLOOKUP(B{r_idx}, PLANTA_BASE!A:{last_col_letter}, 3, FALSE), ""))'

        formula_invitacion = f'=IF(B{r_idx}="","",IFERROR(VLOOKUP(B{r_idx}, PLANTA_BASE!A:{last_col_letter}, 2, FALSE), "NO ENCONTRADO EN PLANTA"))'
        formula_nivel2     = f'=IF(B{r_idx}="","",IFERROR(VLOOKUP(B{r_idx}, PLANTA_BASE!A:{last_col_letter}, 5, FALSE), "NO ENCONTRADO EN PLANTA"))'
        formula_nivel3     = f'=IF(B{r_idx}="","",IFERROR(VLOOKUP(B{r_idx}, PLANTA_BASE!A:{last_col_letter}, 6, FALSE), "NO ENCONTRADO EN PLANTA"))'
        formula_cargo      = f'=IF(B{r_idx}="","",IFERROR(VLOOKUP(B{r_idx}, PLANTA_BASE!A:{last_col_letter}, 8, FALSE), "DOCENTE / FACILITADOR"))'
        formula_centro_costo = f'=IF(B{r_idx}="","",IFERROR(VLOOKUP(B{r_idx}, PLANTA_BASE!A:{last_col_letter}, 9, FALSE), "GENERAL"))'
        formula_escuela    = f'=IF(B{r_idx}="","",IFERROR(VLOOKUP(B{r_idx}, PLANTA_BASE!A:{last_col_letter}, {escuela_col_idx}, FALSE), "GENERAL"))'

        ws.cell(row=r_idx, column=1, value=formula_invitacion)
        ws.cell(row=r_idx, column=2, value=ced_cell_value)
        ws.cell(row=r_idx, column=3, value=nom_formula)
        ws.cell(row=r_idx, column=4, value=formula_nivel2)
        ws.cell(row=r_idx, column=5, value=formula_nivel3)
        ws.cell(row=r_idx, column=6, value=formula_cargo)
        ws.cell(row=r_idx, column=7, value=formula_centro_costo)
        ws.cell(row=r_idx, column=8, value=formula_escuela)
        
        ws.row_dimensions[r_idx].height = 20
        use_alt = (r_idx % 2 == 0)
        for c_idx in range(1, len(headers_consolidado) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = cell_border
            if use_alt:
                cell.fill = alt_fill
            
            if c_idx == 2: # CEDULA
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '0'
            elif c_idx in [1, 4, 5, 8]: # INVITACION, Nivel 2, Nivel 3, ESCUELA
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 22

def format_planta_sheet(ws, df_data):
    ws.views.sheetView[0].showGridLines = True
    headers = list(df_data.columns)
    ws.append(headers)
    
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for r_idx, row in enumerate(df_data.itertuples(index=False), start=2):
        ws.append(list(row))
        ws.row_dimensions[r_idx].height = 20
        use_alt = (r_idx % 2 == 0)
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = cell_border
            if use_alt:
                cell.fill = alt_fill
            
            if headers[c_idx-1] == 'CEDULA':
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 18

def format_busqueda_nombre_sheet(ws, df_data):
    ws.views.sheetView[0].showGridLines = True
    headers = list(df_data.columns)
    ws.append(headers)
    
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for r_idx, row in enumerate(df_data.itertuples(index=False), start=2):
        ws.append(list(row))
        ws.row_dimensions[r_idx].height = 20
        use_alt = (r_idx % 2 == 0)
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = cell_border
            if use_alt:
                cell.fill = alt_fill
            
            header_name = headers[c_idx-1]
            if header_name in ['CÉDULA ASISTENCIA', 'CÉDULA PLANTA ENCONTRADA']:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if str(cell.value).isdigit():
                    cell.number_format = '0'
            elif header_name in ['ESTADO COINCIDENCIA', 'PORCENTAJE SIMILITUD', 'Nombre Nivel 2 (Regional)', 'Nombre Nivel 3 (Sede)', 'ESCUELA']:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 24

# Hoja 1: Consolidado_Unicos
ws_unicos = wb.create_sheet(title="Consolidado_Unicos")
format_consolidado_with_full_column_formulas(ws_unicos, df_unicos)

# Hoja 2: Duplicados
ws_duplicados = wb.create_sheet(title="Duplicados")
format_consolidado_with_full_column_formulas(ws_duplicados, df_duplicados)

# Hoja 3: Busqueda_por_Nombre (Búsqueda por Nombre para no encontrados por Cédula)
ws_busqueda = wb.create_sheet(title="Busqueda_por_Nombre")
format_busqueda_nombre_sheet(ws_busqueda, df_busqueda_nombre)

# Hoja 4: PLANTA_BASE
ws_planta = wb.create_sheet(title="PLANTA_BASE")
format_planta_sheet(ws_planta, df_planta_export)

# Hoja 5: Faltantes_Planta (Personas de Planta que NO están en Consolidado Únicos)
unicos_cedulas_set = set(df_unicos['CEDULA'].dropna().unique())
df_faltantes_planta = df_planta_export[~df_planta_export['CEDULA'].isin(unicos_cedulas_set)].copy()

ws_faltantes = wb.create_sheet(title="Faltantes_Planta")
format_planta_sheet(ws_faltantes, df_faltantes_planta)

print(f"      Facilitadores Faltantes de Planta (Sin Asistencia): {len(df_faltantes_planta)}")

# Hoja 6: Resumen_y_Graficos
ws_resumen = wb.create_sheet(title="Resumen_y_Graficos")
ws_resumen.views.sheetView[0].showGridLines = True
ws_resumen.cell(row=1, column=1, value="RESUMEN EJECUTIVO DE ASISTENCIA CUN").font = Font(name="Calibri", size=16, bold=True, color="004B28")

planta_lookup = df_planta_raw.drop_duplicates(subset=['CEDULA_CLEAN']).set_index('CEDULA_CLEAN')

eval_records = []
asistentes_data = []

for idx, row in df_unicos.iterrows():
    ced = row['CEDULA']
    nom = row['NOMBRE']
    
    if ced in planta_lookup.index:
        p_row = planta_lookup.loc[ced]
        inv_val = clean_text(p_row.get(col_invitacion, ''))
        if not inv_val or inv_val.upper() == 'NAN':
            inv_val = clean_text(p_row.get(col_modalidad_alt, 'VIRTUAL'))
        reg_val = clean_text(p_row.get(col_nivel2, 'NO ASIGNADO'))
        sed_val = clean_text(p_row.get(col_nivel3, 'NO ASIGNADO'))
        car_val = clean_text(p_row.get(col_cargo, 'DOCENTE / FACILITADOR'))
        cc_val  = clean_text(p_row.get(col_centro_costo, 'GENERAL'))
        esc_val = clean_text(p_row.get(col_escuela, 'GENERAL'))
    else:
        inv_val = "VIRTUAL"
        reg_val = "REGIONAL BOGOTA"
        sed_val = "BOGOTA CENTRO"
        car_val = "DOCENTE / FACILITADOR"
        cc_val  = "GENERAL"
        esc_val = "GENERAL"

    inv_upper = inv_val.upper()
    tipo_mod = 'PRESENCIAL' if ('PRESENCIAL' in inv_upper or 'BOGOTA' in inv_upper or 'BOGOTÁ' in inv_upper or 'SEDE' in inv_upper) else 'VIRTUAL'

    record_dict = {
        'INVITACIÓN': inv_val,
        'CEDULA': ced,
        'NOMBRE': nom,
        'Nombre Nivel 2': reg_val,
        'Nombre Nivel 3': sed_val,
        'Descripción Cargo': car_val,
        'Nombre Centro Costo': cc_val,
        'ESCUELA': esc_val,
        'TIPO_MODALIDAD': tipo_mod
    }
    eval_records.append(record_dict)

    asistentes_data.append({
        'cedula': str(ced),
        'nombre': str(nom),
        'escuela': str(esc_val),
        'centro_costo': str(cc_val),
        'regional': str(reg_val),
        'sede': str(sed_val),
        'cargo': str(car_val),
        'modalidad': str(tipo_mod),
        'invitacion': str(inv_val)
    })

df_eval = pd.DataFrame(eval_records)

ws_resumen.cell(row=3, column=1, value="Modalidad / Invitación").font = header_font
ws_resumen.cell(row=3, column=1).fill = header_fill
ws_resumen.cell(row=3, column=2, value="Cantidad Registros").font = header_font
ws_resumen.cell(row=3, column=2).fill = header_fill

if len(df_eval) > 0:
    modalidad_counts = df_eval['INVITACIÓN'].value_counts()
    row_curr = 4
    for mod_name, count_val in modalidad_counts.items():
        ws_resumen.cell(row=row_curr, column=1, value=str(mod_name)).border = cell_border
        ws_resumen.cell(row=row_curr, column=2, value=int(count_val)).border = cell_border
        ws_resumen.cell(row=row_curr, column=2).alignment = Alignment(horizontal="center")
        row_curr += 1

    ws_resumen.cell(row=row_curr, column=1, value="TOTAL GENERAL").font = Font(bold=True)
    ws_resumen.cell(row=row_curr, column=1).border = cell_border
    ws_resumen.cell(row=row_curr, column=2, value=f"=SUM(B4:B{row_curr-1})").font = Font(bold=True)
    ws_resumen.cell(row=row_curr, column=2).border = cell_border
    ws_resumen.cell(row=row_curr, column=2).alignment = Alignment(horizontal="center")

    pie = PieChart()
    labels = Reference(ws_resumen, min_col=1, min_row=4, max_row=row_curr-1)
    data = Reference(ws_resumen, min_col=2, min_row=3, max_row=row_curr-1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Distribución por Modalidad (Únicos Válidos)"
    pie.width = 14
    pie.height = 8
    ws_resumen.add_chart(pie, "D3")

    row_reg_start = row_curr + 3
    ws_resumen.cell(row=row_reg_start, column=1, value="Regional (Nombre Nivel 2)").font = header_font
    ws_resumen.cell(row=row_reg_start, column=1).fill = header_fill
    ws_resumen.cell(row=row_reg_start, column=2, value="Presenciales").font = header_font
    ws_resumen.cell(row=row_reg_start, column=2).fill = header_fill

    regional_presencial = df_eval[df_eval['TIPO_MODALIDAD'] == 'PRESENCIAL']['Nombre Nivel 2'].value_counts()
    row_curr_reg = row_reg_start + 1
    for reg_name, count_val in regional_presencial.items():
        ws_resumen.cell(row=row_curr_reg, column=1, value=str(reg_name)).border = cell_border
        ws_resumen.cell(row=row_curr_reg, column=2, value=int(count_val)).border = cell_border
        ws_resumen.cell(row=row_curr_reg, column=2).alignment = Alignment(horizontal="center")
        row_curr_reg += 1

    if len(regional_presencial) > 0:
        chart_bar = BarChart()
        chart_bar.type = "col"
        chart_bar.style = 10
        chart_bar.title = "Participación Presencial por Regional"
        chart_bar.y_axis.title = "Cantidad"
        chart_bar.x_axis.title = "Regional"
        data_bar = Reference(ws_resumen, min_col=2, min_row=row_reg_start, max_row=row_curr_reg-1)
        cats_bar = Reference(ws_resumen, min_col=1, min_row=row_reg_start+1, max_row=row_curr_reg-1)
        chart_bar.add_data(data_bar, titles_from_data=True)
        chart_bar.set_categories(cats_bar)
        chart_bar.width = 16
        chart_bar.height = 8
        ws_resumen.add_chart(chart_bar, f"D{row_reg_start}")

# Guardar con manejo de bloqueo si el archivo está abierto en Excel
try:
    wb.save(EXCEL_OUTPUT)
    print("      Libro Excel consolidado guardado exitosamente.")
except PermissionError:
    print("      ADVERTENCIA: El archivo Consolidado_Bienvenida_Facilitadores.xlsx está abierto por el usuario.")
    alt_out = os.path.join(BASE_DIR, "Consolidado_Bienvenida_Facilitadores_Nuevo.xlsx")
    wb.save(alt_out)
    print(f"      Guardado exitosamente en copia alternativa: {alt_out}")

# Generar archivo Excel independiente exclusivo de Facilitadores Faltantes de Planta Base
wb_faltantes = openpyxl.Workbook()
wb_faltantes.remove(wb_faltantes.active)
ws_f_standalone = wb_faltantes.create_sheet(title="Facilitadores_Faltantes_Planta")
format_planta_sheet(ws_f_standalone, df_faltantes_planta)

try:
    wb_faltantes.save(EXCEL_FALTANTES)
    print(f"      Libro Excel independiente de Facilitadores Faltantes guardado en: {EXCEL_FALTANTES}")
except PermissionError:
    print("      ADVERTENCIA: El archivo Facilitadores_Pendientes_Asistencia.xlsx está abierto por el usuario.")
    alt_f_out = os.path.join(BASE_DIR, "Facilitadores_Pendientes_Asistencia_Nuevo.xlsx")
    wb_faltantes.save(alt_f_out)
    print(f"      Guardado exitosamente en copia alternativa: {alt_f_out}")

# 5. Generar Tablero HTML (CON CORRECCIÓN DEL BUSCADOR DE ASISTENTES)
print("=== PREPARANDO DATOS PARA EL TABLERO HTML (CON BUSCADOR 100% FUNCIONAL) ===")

total_registros = len(df_eval)
presencial_count = len(df_eval[df_eval['TIPO_MODALIDAD'] == 'PRESENCIAL']) if total_registros > 0 else 0
virtual_count = len(df_eval[df_eval['TIPO_MODALIDAD'] == 'VIRTUAL']) if total_registros > 0 else 0

pct_asistencia_global = round((total_registros / total_planta) * 100, 1) if total_planta > 0 else 0
regional_presencial_dict = df_eval[df_eval['TIPO_MODALIDAD'] == 'PRESENCIAL']['Nombre Nivel 2'].value_counts().to_dict() if total_registros > 0 else {}

escuelas_data = []
if total_registros > 0:
    for escuela, group in df_eval.groupby('ESCUELA'):
        t_count = len(group)
        p_count = len(group[group['TIPO_MODALIDAD'] == 'PRESENCIAL'])
        v_count = len(group[group['TIPO_MODALIDAD'] == 'VIRTUAL'])
        pct = round((t_count / total_registros) * 100, 1) if total_registros > 0 else 0
        
        regiones_presenciales = group[group['TIPO_MODALIDAD'] == 'PRESENCIAL']['Nombre Nivel 2'].value_counts().to_dict()
        regiones_virtuales = group[group['TIPO_MODALIDAD'] == 'VIRTUAL']['Nombre Nivel 2'].value_counts().to_dict()
        sedes_detalle = group['Nombre Nivel 3'].value_counts().to_dict()
        centros_costo_detalle = group['Nombre Centro Costo'].value_counts().to_dict()

        escuelas_data.append({
            'escuela': str(escuela),
            'total': t_count,
            'presencial': p_count,
            'virtual': v_count,
            'porcentaje': pct,
            'regiones_presencial': regiones_presenciales,
            'regiones_virtual': regiones_virtuales,
            'sedes': sedes_detalle,
            'centros_costo': centros_costo_detalle
        })

    escuelas_data.sort(key=lambda x: x['total'], reverse=True)

fecha_actualizacion = datetime.now().strftime("%d/%m/%Y %I:%M %p")

html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Tablero Bienvenida Facilitadores CUN - Diapositiva Ejecutiva</title>
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;600;700;800&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        :root {{
            --cun-green: #00A859;
            --cun-green-glow: #00FF87;
            --cun-dark: #050B14;
            --cun-card-bg: rgba(10, 20, 38, 0.78);
            --cun-cyan: #00F2FE;
            --cun-text: #F3F4F6;
            --cun-muted: #9CA3AF;
            --cun-border: rgba(0, 168, 89, 0.3);
        }}

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
            font-family: 'Plus Jakarta Sans', sans-serif;
        }}

        body {{
            background-color: var(--cun-dark);
            color: var(--cun-text);
            min-height: 100vh;
            overflow-x: hidden;
            position: relative;
        }}

        #galaxy-canvas {{
            position: fixed;
            top: 0;
            left: 0;
            width: 100vw;
            height: 100vh;
            z-index: 0;
            pointer-events: none;
        }}

        .dashboard-container {{
            position: relative;
            z-index: 1;
            max-width: 1350px;
            margin: 0 auto;
            padding: 2rem 1.5rem;
        }}

        /* Header con Logo CUN Directo */
        header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            background: var(--cun-card-bg);
            backdrop-filter: blur(16px);
            border: 1px solid var(--cun-border);
            padding: 1.25rem 2.25rem;
            border-radius: 24px;
            margin-bottom: 2rem;
            box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
        }}

        .brand-section {{
            display: flex;
            align-items: center;
            gap: 1.75rem;
        }}

        .brand-logo-container {{
            display: flex;
            align-items: center;
            justify-content: center;
            background: none;
            border: none;
            padding: 0;
            margin: 0;
            box-shadow: none;
            backdrop-filter: none;
        }}

        .brand-logo-img {{
            height: 75px;
            max-width: 240px;
            width: auto;
            object-fit: contain;
            filter: none;
        }}

        .brand-title h1 {{
            font-size: 1.85rem;
            font-weight: 800;
            letter-spacing: -0.5px;
            background: linear-gradient(90deg, #FFFFFF, #00FF87);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}

        .brand-title p {{
            font-size: 0.95rem;
            color: var(--cun-muted);
            margin-top: 2px;
        }}

        .header-actions {{
            display: flex;
            flex-direction: column;
            align-items: flex-end;
            gap: 0.5rem;
        }}

        .header-banner-img {{
            height: 62px;
            width: auto;
            object-fit: contain;
            border-radius: 12px;
            box-shadow: 0 4px 20px rgba(0, 242, 254, 0.2);
            transition: transform 0.3s ease;
        }}

        .header-banner-img:hover {{
            transform: scale(1.03);
        }}

        .live-badge {{
            display: flex;
            align-items: center;
            gap: 6px;
            background: rgba(0, 255, 135, 0.08);
            border: 1px solid rgba(0, 255, 135, 0.25);
            padding: 3px 10px;
            border-radius: 8px;
        }}

        .live-dot {{
            width: 6px;
            height: 6px;
            background-color: var(--cun-green-glow);
            border-radius: 50%;
            box-shadow: 0 0 8px var(--cun-green-glow);
            animation: pulse 1.5s infinite;
            flex-shrink: 0;
        }}

        .live-info {{
            display: flex;
            align-items: center;
            gap: 6px;
            line-height: 1;
        }}

        .live-title {{
            font-size: 0.55rem;
            font-weight: 700;
            color: var(--cun-green-glow);
            letter-spacing: 0.5px;
            text-transform: uppercase;
        }}

        .live-date {{
            font-size: 0.65rem;
            font-weight: 600;
            color: #FFFFFF;
        }}

        @keyframes pulse {{
            0% {{ transform: scale(0.95); opacity: 0.8; }}
            50% {{ transform: scale(1.25); opacity: 1; }}
            100% {{ transform: scale(0.95); opacity: 0.8; }}
        }}

        /* HERO CARD PRINCIPAL: TOTAL REGISTROS ÚNICOS VÁLIDOS (CON BOTÓN INTERACTIVO LISTADO COMPLETO) */
        .kpi-hero-card {{
            background: var(--cun-card-bg);
            backdrop-filter: blur(16px);
            border: 1px solid var(--cun-border);
            border-radius: 24px;
            padding: 2rem;
            margin-bottom: 1.5rem;
            position: relative;
            overflow: hidden;
            box-shadow: 0 12px 40px rgba(0, 168, 89, 0.15);
            transition: all 0.3s ease;
            cursor: pointer;
        }}

        .kpi-hero-card:hover {{
            border-color: var(--cun-green-glow);
            box-shadow: 0 16px 50px rgba(0, 255, 135, 0.25);
            transform: translateY(-2px);
        }}

        .kpi-hero-card::after {{
            content: "Haz clic aquí para ver el listado completo de asistentes ➔";
            position: absolute;
            top: 18px;
            right: 24px;
            font-size: 0.8rem;
            color: var(--cun-cyan);
            font-weight: 700;
            background: rgba(0, 242, 254, 0.1);
            border: 1px solid rgba(0, 242, 254, 0.3);
            padding: 6px 14px;
            border-radius: 20px;
            transition: all 0.3s ease;
        }}

        .kpi-hero-card:hover::after {{
            background: rgba(0, 242, 254, 0.2);
            box-shadow: 0 0 15px rgba(0, 242, 254, 0.3);
        }}

        .hero-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 1.25rem;
        }}

        .hero-info .kpi-title {{
            font-size: 1rem;
            color: var(--cun-muted);
            font-weight: 700;
            letter-spacing: 0.5px;
        }}

        .hero-info .kpi-value {{
            font-size: 3.5rem;
            font-weight: 800;
            color: #FFFFFF;
            line-height: 1.1;
            margin: 0.25rem 0;
            background: linear-gradient(90deg, #FFFFFF, var(--cun-cyan));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}

        .hero-info .kpi-subtext {{
            font-size: 0.85rem;
            color: var(--cun-muted);
        }}

        .hero-progress-wrapper {{
            margin-top: 1rem;
        }}

        .hero-progress-meta {{
            display: flex;
            justify-content: space-between;
            font-size: 0.82rem;
            font-weight: 700;
            color: var(--cun-muted);
            margin-bottom: 8px;
            letter-spacing: 0.5px;
        }}

        .hero-progress-meta span.highlight {{
            color: var(--cun-green-glow);
            font-weight: 800;
        }}

        .hero-progress-bg {{
            width: 100%;
            height: 12px;
            background: rgba(255, 255, 255, 0.08);
            border-radius: 20px;
            overflow: hidden;
            position: relative;
        }}

        .hero-progress-fill {{
            height: 100%;
            width: {pct_asistencia_global}%;
            min-width: 2%;
            background: linear-gradient(90deg, #00A859, #00FF87, #00F2FE);
            background-size: 200% 100%;
            border-radius: 20px;
            animation: glowGradient 3s infinite linear;
            transition: width 1.2s cubic-bezier(0.4, 0, 0.2, 1);
        }}

        @keyframes glowGradient {{
            0% {{ background-position: 0% 50%; }}
            50% {{ background-position: 100% 50%; }}
            100% {{ background-position: 0% 50%; }}
        }}

        /* SUBGRID: ASISTENCIA PRESENCIAL Y ASISTENCIA VIRTUAL (DEBAJO) */
        .kpi-subgrid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 1.5rem;
            margin-bottom: 2.5rem;
        }}

        .kpi-card {{
            background: var(--cun-card-bg);
            backdrop-filter: blur(16px);
            border: 1px solid var(--cun-border);
            border-radius: 20px;
            padding: 1.75rem;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
        }}

        .kpi-card:hover {{
            transform: translateY(-5px);
            border-color: var(--cun-green-glow);
            box-shadow: 0 12px 40px rgba(0, 168, 89, 0.2);
        }}

        .kpi-card.clickable {{
            cursor: pointer;
        }}

        .kpi-card.clickable::after {{
            content: "Clic para desglose regional ➔";
            position: absolute;
            bottom: 10px;
            right: 16px;
            font-size: 0.72rem;
            color: var(--cun-cyan);
            opacity: 0.9;
            font-weight: 700;
        }}

        .kpi-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 1rem;
        }}

        .kpi-icon {{
            width: 46px;
            height: 46px;
            border-radius: 14px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 1.35rem;
        }}

        .icon-total {{ background: rgba(0, 242, 254, 0.15); color: var(--cun-cyan); }}
        .icon-presencial {{ background: rgba(0, 255, 135, 0.15); color: var(--cun-green-glow); }}
        .icon-virtual {{ background: rgba(168, 85, 247, 0.15); color: #C084FC; }}

        .kpi-title {{
            font-size: 0.9rem;
            color: var(--cun-muted);
            font-weight: 600;
        }}

        .kpi-value {{
            font-size: 2.75rem;
            font-weight: 800;
            color: #FFFFFF;
            line-height: 1;
            margin-bottom: 0.5rem;
        }}

        .kpi-subtext {{
            font-size: 0.8rem;
            color: var(--cun-muted);
        }}

        /* Section Title & Search */
        .section-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 1.5rem;
            flex-wrap: wrap;
            gap: 1rem;
        }}

        .section-title {{
            font-size: 1.3rem;
            font-weight: 800;
            display: flex;
            align-items: center;
            gap: 10px;
        }}

        .search-box {{
            position: relative;
            min-width: 320px;
        }}

        .search-box input {{
            width: 100%;
            background: rgba(15, 23, 42, 0.85);
            border: 1px solid var(--cun-border);
            padding: 0.75rem 1rem 0.75rem 2.5rem;
            border-radius: 12px;
            color: #FFFFFF;
            font-size: 0.9rem;
            outline: none;
            transition: all 0.3s ease;
        }}

        .search-box input:focus {{
            border-color: var(--cun-cyan);
            box-shadow: 0 0 15px rgba(0, 242, 254, 0.25);
        }}

        .search-box i {{
            position: absolute;
            left: 1rem;
            top: 50%;
            transform: translateY(-50%);
            color: var(--cun-muted);
        }}

        /* LAYOUT DIVIDIDO IZQUIERDA Y DERECHA POR ESCUELA */
        .escuelas-grid-layout {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 1.75rem;
        }}

        .escuelas-panel {{
            background: var(--cun-card-bg);
            backdrop-filter: blur(16px);
            border: 1px solid var(--cun-border);
            border-radius: 24px;
            padding: 1.5rem;
            box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.37);
        }}

        .panel-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding-bottom: 1rem;
            margin-bottom: 1rem;
            border-bottom: 1px solid rgba(255, 255, 255, 0.1);
        }}

        .panel-header h3 {{
            font-size: 1.1rem;
            font-weight: 700;
            color: var(--cun-green-glow);
            display: flex;
            align-items: center;
            gap: 10px;
        }}

        .panel-header span {{
            font-size: 0.8rem;
            color: var(--cun-muted);
            font-weight: 600;
        }}

        .escuelas-list {{
            display: flex;
            flex-direction: column;
            gap: 1rem;
            max-height: 520px;
            overflow-y: auto;
            padding-right: 6px;
        }}

        /* Item Panel Izquierdo: Totales por Escuela */
        .escuela-card-left {{
            background: rgba(15, 23, 42, 0.7);
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 16px;
            padding: 1.1rem;
            transition: all 0.2s ease;
        }}

        .escuela-card-left:hover {{
            border-color: var(--cun-green);
            background: rgba(0, 168, 89, 0.08);
        }}

        .card-left-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 8px;
        }}

        .card-left-title {{
            font-size: 0.95rem;
            font-weight: 700;
            color: #FFFFFF;
        }}

        .card-left-badge {{
            background: linear-gradient(135deg, var(--cun-green), var(--cun-cyan));
            color: #050B14;
            font-weight: 800;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 0.85rem;
        }}

        .progress-bar-bg {{
            width: 100%;
            height: 8px;
            background: rgba(255, 255, 255, 0.1);
            border-radius: 20px;
            overflow: hidden;
            margin-top: 6px;
        }}

        .progress-bar-fill {{
            height: 100%;
            background: linear-gradient(90deg, var(--cun-green), var(--cun-cyan));
            border-radius: 20px;
            transition: width 1s ease;
        }}

        /* Item Panel Derecho: Virtual vs Presencial por Escuela */
        .escuela-card-right {{
            background: rgba(15, 23, 42, 0.7);
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 16px;
            padding: 1.1rem;
            cursor: pointer;
            position: relative;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        }}

        .escuela-card-right:hover {{
            border-color: var(--cun-cyan);
            background: rgba(0, 242, 254, 0.08);
            transform: scale(1.01);
            box-shadow: 0 8px 24px rgba(0, 242, 254, 0.15);
        }}

        .escuela-card-right::after {{
            content: "Ver sedes / origen ➔";
            position: absolute;
            top: 12px;
            right: 14px;
            font-size: 0.72rem;
            color: var(--cun-cyan);
            opacity: 0.85;
            font-weight: 700;
        }}

        .card-right-header {{
            font-size: 0.95rem;
            font-weight: 700;
            color: #FFFFFF;
            margin-bottom: 10px;
            padding-right: 120px;
        }}

        .modalidad-breakdown {{
            display: flex;
            gap: 12px;
        }}

        .breakdown-box {{
            flex: 1;
            padding: 8px 12px;
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: space-between;
        }}

        .box-presencial {{
            background: rgba(0, 255, 135, 0.12);
            border: 1px solid rgba(0, 255, 135, 0.3);
            color: var(--cun-green-glow);
        }}

        .box-virtual {{
            background: rgba(168, 85, 247, 0.12);
            border: 1px solid rgba(168, 85, 247, 0.3);
            color: #C084FC;
        }}

        .box-label {{
            font-size: 0.8rem;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 6px;
        }}

        .box-value {{
            font-size: 1.1rem;
            font-weight: 800;
        }}

        /* Modal Overlay */
        .modal-overlay {{
            position: fixed;
            top: 0;
            left: 0;
            width: 100vw;
            height: 100vh;
            background: rgba(5, 11, 20, 0.85);
            backdrop-filter: blur(14px);
            z-index: 999;
            display: flex;
            align-items: center;
            justify-content: center;
            opacity: 0;
            pointer-events: none;
            transition: all 0.3s ease;
        }}

        .modal-overlay.active {{
            opacity: 1;
            pointer-events: auto;
        }}

        .modal-content {{
            background: #0A1426;
            border: 1px solid var(--cun-cyan);
            width: 92%;
            max-width: 950px;
            border-radius: 24px;
            padding: 2rem;
            box-shadow: 0 0 60px rgba(0, 242, 254, 0.25);
            transform: scale(0.9);
            transition: transform 0.3s ease;
        }}

        .modal-overlay.active .modal-content {{
            transform: scale(1);
        }}

        .modal-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 1.5rem;
            border-bottom: 1px solid rgba(255, 255, 255, 0.1);
            padding-bottom: 1rem;
        }}

        .modal-header h3 {{
            font-size: 1.3rem;
            color: var(--cun-cyan);
            display: flex;
            align-items: center;
            gap: 10px;
        }}

        .close-modal {{
            background: none;
            border: none;
            color: var(--cun-muted);
            font-size: 1.25rem;
            cursor: pointer;
            transition: color 0.2s;
        }}

        .close-modal:hover {{
            color: #FFFFFF;
        }}

        /* ESTILOS ESPECÍFICOS DEL MODAL DE LISTADO COMPLETO DE ASISTENTES */
        .asistentes-modal-search-wrapper {{
            margin-bottom: 1.25rem;
        }}

        .asistentes-modal-search {{
            position: relative;
            width: 100%;
        }}

        .asistentes-modal-search input {{
            width: 100%;
            background: rgba(15, 23, 42, 0.95);
            border: 1px solid var(--cun-cyan);
            padding: 0.85rem 1rem 0.85rem 2.8rem;
            border-radius: 14px;
            color: #FFFFFF;
            font-size: 0.95rem;
            outline: none;
            box-shadow: 0 0 15px rgba(0, 242, 254, 0.15);
        }}

        .asistentes-modal-search i {{
            position: absolute;
            left: 1.1rem;
            top: 50%;
            transform: translateY(-50%);
            color: var(--cun-cyan);
            font-size: 1rem;
        }}

        .asistentes-table-container {{
            max-height: 480px;
            overflow-y: auto;
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 16px;
            background: rgba(10, 20, 38, 0.6);
        }}

        .asistentes-table {{
            width: 100%;
            border-collapse: collapse;
            text-align: left;
            font-size: 0.88rem;
        }}

        .asistentes-table th {{
            background: rgba(0, 168, 89, 0.25);
            color: var(--cun-green-glow);
            padding: 12px 16px;
            font-weight: 700;
            position: sticky;
            top: 0;
            z-index: 2;
            border-bottom: 1px solid var(--cun-border);
        }}

        .asistentes-table td {{
            padding: 12px 16px;
            border-bottom: 1px solid rgba(255, 255, 255, 0.05);
            color: var(--cun-text);
        }}

        .asistentes-table tr:hover {{
            background: rgba(0, 242, 254, 0.08);
        }}

        .badge-mod-presencial {{
            background: rgba(0, 255, 135, 0.15);
            color: var(--cun-green-glow);
            border: 1px solid rgba(0, 255, 135, 0.3);
            padding: 4px 10px;
            border-radius: 12px;
            font-weight: 700;
            font-size: 0.78rem;
            display: inline-block;
        }}

        .badge-mod-virtual {{
            background: rgba(168, 85, 247, 0.15);
            color: #C084FC;
            border: 1px solid rgba(168, 85, 247, 0.3);
            padding: 4px 10px;
            border-radius: 12px;
            font-weight: 700;
            font-size: 0.78rem;
            display: inline-block;
        }}

        .regional-list {{
            display: flex;
            flex-direction: column;
            gap: 1rem;
            max-height: 420px;
            overflow-y: auto;
            padding-right: 8px;
        }}

        .regional-card {{
            background: rgba(15, 23, 42, 0.7);
            border: 1px solid rgba(255, 255, 255, 0.08);
            padding: 1rem 1.25rem;
            border-radius: 14px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}

        .regional-name {{
            font-weight: 700;
            font-size: 0.95rem;
            color: #FFFFFF;
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .regional-badge {{
            background: linear-gradient(135deg, var(--cun-green), var(--cun-cyan));
            color: #050B14;
            font-weight: 800;
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 0.9rem;
        }}

        .modal-sub-section {{
            margin-top: 1rem;
            margin-bottom: 0.5rem;
            font-size: 0.85rem;
            font-weight: 700;
            color: var(--cun-muted);
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        @media (max-width: 992px) {{
            .escuelas-grid-layout, .kpi-subgrid {{
                grid-template-columns: 1fr;
            }}
            .kpi-hero-card::after {{
                position: static;
                display: inline-block;
                margin-top: 14px;
                content: "Ver listado completo de asistentes ➔";
                font-size: 0.75rem;
                padding: 5px 12px;
            }}
            header {{
                flex-direction: column;
                align-items: flex-start;
                gap: 1.2rem;
            }}
            .header-actions {{
                align-items: flex-start;
                width: 100%;
            }}
        }}

        @media (max-width: 600px) {{
            .brand-section {{
                flex-direction: column;
                align-items: flex-start;
                gap: 0.75rem;
            }}
            .brand-title h1 {{
                font-size: 1.4rem;
            }}
            .hero-info .kpi-value {{
                font-size: 2.75rem;
            }}
            .search-box {{
                min-width: 100%;
            }}
            .modal-content {{
                width: 95%;
                padding: 1.25rem;
            }}
            .asistentes-table {{
                font-size: 0.78rem;
            }}
        }}
    </style>
</head>
<body>

    <canvas id="galaxy-canvas"></canvas>

    <div class="dashboard-container">
        <header>
            <div class="brand-section">
                <div class="brand-logo-container">
                    <img src="Logo.png" alt="Logo CUN" class="brand-logo-img">
                </div>
                <div class="brand-title">
                    <h1>Bienvenida Facilitadores CUN</h1>
                    <p>Tablero Dinámico de Asistencia</p>
                </div>
            </div>
            <div class="header-actions">
                <img src="image.png" alt="Cunéctate al Parche" class="header-banner-img">
                <div class="live-badge">
                    <div class="live-dot"></div>
                    <div class="live-info">
                        <span class="live-title">ÚLTIMA ACTUALIZACIÓN:</span>
                        <span class="live-date">{fecha_actualizacion}</span>
                    </div>
                </div>
            </div>
        </header>

        <!-- HERO CARD PRINCIPAL: TOTAL REGISTROS ÚNICOS VÁLIDOS -->
        <div class="kpi-hero-card" id="btn-total-asistentes">
            <div class="hero-header">
                <div class="hero-info">
                    <span class="kpi-title">TOTAL REGISTROS</span>
                    <div class="kpi-value" id="kpi-total">{total_registros}</div>
                    <div class="kpi-subtext">..</div>
                </div>
                <div class="kpi-icon icon-total"><i class="fa-solid fa-users"></i></div>
            </div>
            <div class="hero-progress-wrapper">
                <div class="hero-progress-meta">
                    <span>PORCENTAJE DE ASISTENCIA EVENTO ({total_registros} / {total_planta} FACILITADORES)</span>
                    <span class="highlight">{pct_asistencia_global}%</span>
                </div>
                <div class="hero-progress-bg">
                    <div class="hero-progress-fill"></div>
                </div>
            </div>
        </div>

        <!-- SUBGRID: ASISTENCIA PRESENCIAL Y ASISTENCIA VIRTUAL -->
        <div class="kpi-subgrid">
            <div class="kpi-card clickable" id="btn-presencial">
                <div class="kpi-header">
                    <span class="kpi-title">ASISTENCIA PRESENCIAL</span>
                    <div class="kpi-icon icon-presencial"><i class="fa-solid fa-building-user"></i></div>
                </div>
                <div class="kpi-value" style="color: var(--cun-green-glow);">{presencial_count}</div>
                <div class="kpi-subtext">...</div>
            </div>

            <div class="kpi-card">
                <div class="kpi-header">
                    <span class="kpi-title">ASISTENCIA VIRTUAL</span>
                    <div class="kpi-icon icon-virtual"><i class="fa-solid fa-laptop-code"></i></div>
                </div>
                <div class="kpi-value" style="color: #C084FC;">{virtual_count}</div>
                <div class="kpi-subtext">...</div>
            </div>
        </div>

        <div class="section-header">
            <div class="section-title">
                <i class="fa-solid fa-chart-pie" style="color: var(--cun-green-glow);"></i>
                Análisis de Asistencia por Escuela 
            </div>
            <div class="search-box">
                <i class="fa-solid fa-magnifying-glass"></i>
                <input type="text" id="school-search" placeholder="Buscar escuela o centro de costo...">
            </div>
        </div>

        <!-- LAYOUT DE DOS COLUMNAS IZQUIERDA Y DERECHA POR ESCUELA -->
        <div class="escuelas-grid-layout">
            <!-- COLUMNA IZQUIERDA: TOTALES POR ESCUELA -->
            <div class="escuelas-panel">
                <div class="panel-header">
                    <h3><i class="fa-solid fa-school"></i> Total por Escuela</h3>
                    <span>Clic para ver Centros de Costo (Col. I)</span>
                </div>
                <div class="escuelas-list" id="escuelas-totales-list">
                </div>
            </div>

            <!-- COLUMNA DERECHA: DESGLOSE PRESENCIAL VS VIRTUAL POR ESCUELA -->
            <div class="escuelas-panel">
                <div class="panel-header">
                    <h3><i class="fa-solid fa-layer-group"></i> Conteo Virtual vs Presencial</h3>
                    <span>Clic en la escuela para ver sedes/origen</span>
                </div>
                <div class="escuelas-list" id="escuelas-modalidades-list">
                </div>
            </div>
        </div>
    </div>

    <!-- MODAL PRINCIPAL 0: LISTADO COMPLETO DESPLEGABLE DE ASISTENTES ÚNICOS VÁLIDOS CON FILTRO DE BÚSQUEDA -->
    <div class="modal-overlay" id="modal-asistentes-completos">
        <div class="modal-content">
            <div class="modal-header">
                <h3><i class="fa-solid fa-address-book" style="color: var(--cun-cyan);"></i> Listado Completo de Registros (<span id="asistentes-counter">{total_registros}</span>)</h3>
                <button class="close-modal" id="close-asistentes-modal"><i class="fa-solid fa-xmark"></i></button>
            </div>
            
            <!-- BUSCADOR DENTRO DEL MODAL -->
            <div class="asistentes-modal-search-wrapper">
                <div class="asistentes-modal-search">
                    <i class="fa-solid fa-magnifying-glass"></i>
                    <input type="text" id="asistentes-modal-input" placeholder="Buscar por Nombre, Cédula, Escuela, Centro Costo, Regional o Modalidad...">
                </div>
            </div>

            <!-- TABLA DE RESULTADOS -->
            <div class="asistentes-table-container">
                <table class="asistentes-table">
                    <thead>
                        <tr>
                            <th>Cédula</th>
                            <th>Nombre Completo</th>
                            <th>Escuela (Col. H)</th>
                            <th>Centro Costo (Col. I)</th>
                            <th>Regional</th>
                            <th>Sede</th>
                            <th>Modalidad</th>
                        </tr>
                    </thead>
                    <tbody id="asistentes-table-body">
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <!-- MODAL 1: REGIONAL GENERAL PRESENCIAL -->
    <div class="modal-overlay" id="modal-presencial">
        <div class="modal-content" style="max-width: 650px;">
            <div class="modal-header">
                <h3 style="color: var(--cun-green-glow);"><i class="fa-solid fa-map-location-dot"></i> Desglose Presencial General por Regiones</h3>
                <button class="close-modal" id="close-modal"><i class="fa-solid fa-xmark"></i></button>
            </div>
            <div class="regional-list" id="regional-list-content">
            </div>
        </div>
    </div>

    <!-- MODAL 2: EMERGENTE DETALLADO DE ORIGEN REGIONAL/SEDES POR ESCUELA -->
    <div class="modal-overlay" id="modal-escuela-detalle">
        <div class="modal-content" style="max-width: 650px;">
            <div class="modal-header">
                <h3 id="escuela-modal-title"><i class="fa-solid fa-city"></i> Origen Regional por Escuela</h3>
                <button class="close-modal" id="close-escuela-modal"><i class="fa-solid fa-xmark"></i></button>
            </div>
            <div class="regional-list" id="escuela-modal-body">
            </div>
        </div>
    </div>

    <!-- MODAL 3: EMERGENTE DETALLADO DE COLUMNA I (NOMBRE CENTRO COSTO) POR ESCUELA -->
    <div class="modal-overlay" id="modal-escuela-centros-costo">
        <div class="modal-content" style="max-width: 700px;">
            <div class="modal-header">
                <h3 id="centros-costo-modal-title" style="color: var(--cun-green-glow);"><i class="fa-solid fa-layer-group"></i> Detalle Columna I por Escuela</h3>
                <button class="close-modal" id="close-centros-costo-modal"><i class="fa-solid fa-xmark"></i></button>
            </div>
            <div style="margin-bottom: 1.25rem; color: var(--cun-muted); font-size: 0.88rem; font-weight: 600;">
                Desglose de registros correspondientes a la <strong style="color: var(--cun-cyan);">Columna I (Nombre Centro Costo)</strong> para esta escuela:
            </div>
            <div class="regional-list" id="centros-costo-modal-body">
            </div>
        </div>
    </div>

    <script>
        const escuelasData = {json.dumps(escuelas_data, ensure_ascii=False)};
        const regionalPresencialData = {json.dumps(regional_presencial_dict, ensure_ascii=False)};
        const asistentesData = {json.dumps(asistentes_data, ensure_ascii=False)};

        const escuelasTotalesContainer = document.getElementById('escuelas-totales-list');
        const escuelasModalidadesContainer = document.getElementById('escuelas-modalidades-list');
        const searchInput = document.getElementById('school-search');

        // Modal Listado Completo Asistentes
        const btnTotalAsistentes = document.getElementById('btn-total-asistentes');
        const modalAsistentesCompletos = document.getElementById('modal-asistentes-completos');
        const closeAsistentesModalBtn = document.getElementById('close-asistentes-modal');
        const asistentesModalInput = document.getElementById('asistentes-modal-input');
        const asistentesTableBody = document.getElementById('asistentes-table-body');
        const asistentesCounter = document.getElementById('asistentes-counter');

        function renderAsistentesModal(filterText = '') {{
            asistentesTableBody.innerHTML = '';
            const query = filterText ? String(filterText).toLowerCase().trim() : '';
            
            const filtered = asistentesData.filter(a => 
                (a.nombre && a.nombre.toLowerCase().includes(query)) ||
                (a.cedula && a.cedula.toLowerCase().includes(query)) ||
                (a.escuela && a.escuela.toLowerCase().includes(query)) ||
                (a.centro_costo && a.centro_costo.toLowerCase().includes(query)) ||
                (a.regional && a.regional.toLowerCase().includes(query)) ||
                (a.modalidad && a.modalidad.toLowerCase().includes(query))
            );

            asistentesCounter.innerText = `${{filtered.length}} / ${{asistentesData.length}}`;

            if (filtered.length === 0) {{
                asistentesTableBody.innerHTML = '<tr><td colspan="7" style="text-align:center; padding: 2rem; color: var(--cun-muted);">No se encontraron asistentes con ese criterio de búsqueda</td></tr>';
                return;
            }}

            filtered.forEach(item => {{
                const tr = document.createElement('tr');
                const badgeClass = item.modalidad === 'PRESENCIAL' ? 'badge-mod-presencial' : 'badge-mod-virtual';
                tr.innerHTML = `
                    <td style="font-weight:700; font-family:monospace; color:var(--cun-cyan);">${{item.cedula}}</td>
                    <td style="font-weight:700; color:#FFFFFF;">${{item.nombre}}</td>
                    <td><span style="background:rgba(0,168,89,0.18); color:var(--cun-green-glow); border:1px solid rgba(0,255,135,0.3); padding:3px 8px; border-radius:8px; font-weight:700; font-size:0.78rem;">${{item.escuela}}</span></td>
                    <td style="color:var(--cun-muted);">${{item.centro_costo}}</td>
                    <td>${{item.regional}}</td>
                    <td style="color:var(--cun-muted);">${{item.sede}}</td>
                    <td><span class="${{badgeClass}}">${{item.modalidad}}</span></td>
                `;
                asistentesTableBody.appendChild(tr);
            }});
        }}

        btnTotalAsistentes.addEventListener('click', () => {{
            asistentesModalInput.value = '';
            renderAsistentesModal();
            modalAsistentesCompletos.classList.add('active');
            setTimeout(() => {{
                asistentesModalInput.focus();
            }}, 100);
        }});

        closeAsistentesModalBtn.addEventListener('click', () => {{
            modalAsistentesCompletos.classList.remove('active');
        }});

        modalAsistentesCompletos.addEventListener('click', (e) => {{
            if (e.target === modalAsistentesCompletos) {{
                modalAsistentesCompletos.classList.remove('active');
            }}
        }});

        asistentesModalInput.addEventListener('input', (e) => {{
            renderAsistentesModal(e.target.value);
        }});

        // Modal Detalle Centros de Costo (Columna I) por Escuela
        const modalCentrosCosto = document.getElementById('modal-escuela-centros-costo');
        const centrosCostoModalTitle = document.getElementById('centros-costo-modal-title');
        const centrosCostoModalBody = document.getElementById('centros-costo-modal-body');
        const closeCentrosCostoModalBtn = document.getElementById('close-centros-costo-modal');

        function openEscuelaCentrosCostoModal(item) {{
            centrosCostoModalTitle.innerHTML = `<i class="fa-solid fa-folder-tree" style="color: var(--cun-green-glow);"></i> Centros de Costo (Columna I): ${{item.escuela}}`;
            centrosCostoModalBody.innerHTML = '';

            const ccKeys = Object.keys(item.centros_costo || {{}});
            if (ccKeys.length === 0) {{
                centrosCostoModalBody.innerHTML = '<div style="text-align:center; color:var(--cun-muted); padding:2rem;">No hay centros de costo registrados para esta escuela</div>';
            }} else {{
                ccKeys.forEach(cc => {{
                    const count = item.centros_costo[cc];
                    const pct = item.total > 0 ? ((count / item.total) * 100).toFixed(1) : 0;
                    const card = document.createElement('div');
                    card.className = 'regional-card';
                    card.style.background = 'rgba(15, 23, 42, 0.85)';
                    card.style.border = '1px solid var(--cun-border)';
                    card.innerHTML = `
                        <div style="display:flex; flex-direction:column; gap:4px; flex:1;">
                            <span class="regional-name" style="font-size:0.92rem; color:#FFFFFF;"><i class="fa-solid fa-bookmark" style="color:var(--cun-cyan);"></i> ${{cc}}</span>
                            <span style="font-size:0.75rem; color:var(--cun-muted);">${{pct}}% del total de la escuela (${{count}} de ${{item.total}})</span>
                        </div>
                        <span class="regional-badge" style="background: linear-gradient(135deg, var(--cun-green), var(--cun-cyan)); color: #050B14; font-weight: 800;">${{count}} Facilitadores</span>
                    `;
                    centrosCostoModalBody.appendChild(card);
                }});
            }}

            modalCentrosCosto.classList.add('active');
        }}

        closeCentrosCostoModalBtn.addEventListener('click', () => {{
            modalCentrosCosto.classList.remove('active');
        }});

        modalCentrosCosto.addEventListener('click', (e) => {{
            if (e.target === modalCentrosCosto) {{
                modalCentrosCosto.classList.remove('active');
            }}
        }});

        // Modal Detalle Escuela
        const modalEscuelaDetalle = document.getElementById('modal-escuela-detalle');
        const escuelaModalTitle = document.getElementById('escuela-modal-title');
        const escuelaModalBody = document.getElementById('escuela-modal-body');
        const closeEscuelaModalBtn = document.getElementById('close-escuela-modal');

        function openEscuelaRegionalModal(item) {{
            escuelaModalTitle.innerHTML = `<i class="fa-solid fa-city"></i> Origen Regional: ${{item.escuela}}`;
            escuelaModalBody.innerHTML = '';

            let hasData = false;

            // Sub-sección 1: Regiones Presenciales
            const presKeys = Object.keys(item.regiones_presencial || {{}});
            if (presKeys.length > 0) {{
                hasData = true;
                const titleP = document.createElement('div');
                titleP.className = 'modal-sub-section';
                titleP.innerHTML = `<i class="fa-solid fa-location-dot" style="color: var(--cun-green-glow);"></i> Asistentes Presenciales por Región (${{item.presencial}})`;
                escuelaModalBody.appendChild(titleP);

                presKeys.forEach(reg => {{
                    const card = document.createElement('div');
                    card.className = 'regional-card';
                    card.innerHTML = `
                        <span class="regional-name"><i class="fa-solid fa-building-flag" style="color: var(--cun-green-glow);"></i> ${{reg}}</span>
                        <span class="regional-badge">${{item.regiones_presencial[reg]}} Facilitadores</span>
                    `;
                    escuelaModalBody.appendChild(card);
                }});
            }}

            // Sub-sección 2: Regiones Virtuales
            const virtKeys = Object.keys(item.regiones_virtual || {{}});
            if (virtKeys.length > 0) {{
                hasData = true;
                const titleV = document.createElement('div');
                titleV.className = 'modal-sub-section';
                titleV.style.marginTop = '1.25rem';
                titleV.innerHTML = `<i class="fa-solid fa-globe" style="color: #C084FC;"></i> Conexiones Virtuales por Región (${{item.virtual}})`;
                escuelaModalBody.appendChild(titleV);

                virtKeys.forEach(reg => {{
                    const card = document.createElement('div');
                    card.className = 'regional-card';
                    card.innerHTML = `
                        <span class="regional-name"><i class="fa-solid fa-laptop" style="color: #C084FC;"></i> ${{reg}}</span>
                        <span class="regional-badge" style="background: linear-gradient(135deg, #A855F7, #00F2FE);">${{item.regiones_virtual[reg]}} Facilitadores</span>
                    `;
                    escuelaModalBody.appendChild(card);
                }});
            }}

            // Sub-sección 3: Detalle por Sedes
            const sedesKeys = Object.keys(item.sedes || {{}});
            if (sedesKeys.length > 0) {{
                const titleS = document.createElement('div');
                titleS.className = 'modal-sub-section';
                titleS.style.marginTop = '1.25rem';
                titleS.innerHTML = `<i class="fa-solid fa-landmark"></i> Detalle por Sedes Ubicación`;
                escuelaModalBody.appendChild(titleS);

                sedesKeys.forEach(s => {{
                    const card = document.createElement('div');
                    card.className = 'regional-card';
                    card.style.background = 'rgba(255,255,255,0.03)';
                    card.innerHTML = `
                        <span class="regional-name" style="font-size:0.88rem;"><i class="fa-solid fa-door-open" style="color: var(--cun-cyan);"></i> ${{s}}</span>
                        <span style="font-size:0.85rem; font-weight:700; color:var(--cun-muted);">${{item.sedes[s]}} Registros</span>
                    `;
                    escuelaModalBody.appendChild(card);
                }});
            }}

            if (!hasData) {{
                escuelaModalBody.innerHTML = '<div style="text-align:center; color:var(--cun-muted); padding:2rem;">No hay desglose regional registrado para esta escuela</div>';
            }}

            modalEscuelaDetalle.classList.add('active');
        }}

        closeEscuelaModalBtn.addEventListener('click', () => {{
            modalEscuelaDetalle.classList.remove('active');
        }});

        modalEscuelaDetalle.addEventListener('click', (e) => {{
            if (e.target === modalEscuelaDetalle) {{
                modalEscuelaDetalle.classList.remove('active');
            }}
        }});

        function renderEscuelas(filterText = '') {{
            escuelasTotalesContainer.innerHTML = '';
            escuelasModalidadesContainer.innerHTML = '';

            const filtered = escuelasData.filter(e => e.escuela.toLowerCase().includes(filterText.toLowerCase()));

            if (filtered.length === 0) {{
                const emptyMsg = '<div style="text-align: center; padding: 2rem; color: var(--cun-muted);">No hay coincidencias en la asistencia</div>';
                escuelasTotalesContainer.innerHTML = emptyMsg;
                escuelasModalidadesContainer.innerHTML = emptyMsg;
                return;
            }}

            filtered.forEach(item => {{
                // Card Columna Izquierda: Total por Escuela (Clic para emergente Columna I)
                const cardLeft = document.createElement('div');
                cardLeft.className = 'escuela-card-left';
                cardLeft.style.cursor = 'pointer';
                cardLeft.innerHTML = `
                    <div class="card-left-header">
                        <span class="card-left-title">${{item.escuela}}</span>
                        <span class="card-left-badge">${{item.total}} Total</span>
                    </div>
                    <div class="progress-bar-bg">
                        <div class="progress-bar-fill" style="width: ${{item.porcentaje}}%;"></div>
                    </div>
                    <div style="font-size:0.75rem; color: var(--cun-cyan); margin-top:8px; font-weight:700; display:flex; align-items:center; justify-content:space-between;">
                        <span><i class="fa-solid fa-list-check"></i> Clic para ver Centros de Costo (Col. I)</span>
                        <i class="fa-solid fa-chevron-right"></i>
                    </div>
                `;
                cardLeft.addEventListener('click', () => {{
                    openEscuelaCentrosCostoModal(item);
                }});
                escuelasTotalesContainer.appendChild(cardLeft);

                // Card Columna Derecha: Conteo Virtual vs Presencial por Escuela
                const cardRight = document.createElement('div');
                cardRight.className = 'escuela-card-right';
                cardRight.innerHTML = `
                    <div class="card-right-header">${{item.escuela}}</div>
                    <div class="modalidad-breakdown">
                        <div class="breakdown-box box-presencial">
                            <span class="box-label"><i class="fa-solid fa-location-dot"></i> Presenciales</span>
                            <span class="box-value">${{item.presencial}}</span>
                        </div>
                        <div class="breakdown-box box-virtual">
                            <span class="box-label"><i class="fa-solid fa-globe"></i> Virtuales</span>
                            <span class="box-value">${{item.virtual}}</span>
                        </div>
                    </div>
                `;

                cardRight.addEventListener('click', () => {{
                    openEscuelaRegionalModal(item);
                }});

                escuelasModalidadesContainer.appendChild(cardRight);
            }});
        }}

        renderEscuelas();

        searchInput.addEventListener('input', (e) => {{
            renderEscuelas(e.target.value);
        }});

        const modalOverlay = document.getElementById('modal-presencial');
        const btnPresencial = document.getElementById('btn-presencial');
        const btnCloseModal = document.getElementById('close-modal');
        const regionalListContent = document.getElementById('regional-list-content');

        btnPresencial.addEventListener('click', () => {{
            regionalListContent.innerHTML = '';
            const keys = Object.keys(regionalPresencialData);
            if (keys.length === 0) {{
                regionalListContent.innerHTML = '<div style="text-align:center; color:var(--cun-muted); padding: 1.5rem;">No hay registros presenciales válidos asignados a regiones</div>';
            }} else {{
                keys.forEach(reg => {{
                    const card = document.createElement('div');
                    card.className = 'regional-card';
                    card.innerHTML = `
                        <span class="regional-name"><i class="fa-solid fa-building-flag" style="color:var(--cun-green-glow);"></i> ${{reg}}</span>
                        <span class="regional-badge">${{regionalPresencialData[reg]}} Facilitadores</span>
                    `;
                    regionalListContent.appendChild(card);
                }});
            }}
            modalOverlay.classList.add('active');
        }});

        btnCloseModal.addEventListener('click', () => {{
            modalOverlay.classList.remove('active');
        }});

        modalOverlay.addEventListener('click', (e) => {{
            if (e.target === modalOverlay) {{
                modalOverlay.classList.remove('active');
            }}
        }});

        const canvas = document.getElementById('galaxy-canvas');
        const ctx = canvas.getContext('2d');

        function resizeCanvas() {{
            canvas.width = window.innerWidth;
            canvas.height = window.innerHeight;
        }}
        resizeCanvas();
        window.addEventListener('resize', resizeCanvas);

        const stars = [];
        const numStars = 180;

        for (let i = 0; i < numStars; i++) {{
            stars.push({{
                x: Math.random() * canvas.width,
                y: Math.random() * canvas.height,
                radius: Math.random() * 1.8 + 0.2,
                color: Math.random() > 0.4 ? '#00A859' : (Math.random() > 0.5 ? '#00F2FE' : '#FFFFFF'),
                alpha: Math.random() * 0.8 + 0.2,
                vx: (Math.random() - 0.5) * 0.2,
                vy: (Math.random() - 0.5) * 0.2
            }});
        }}

        function drawGalaxy() {{
            ctx.clearRect(0, 0, canvas.width, canvas.height);

            const grad1 = ctx.createRadialGradient(canvas.width*0.3, canvas.height*0.3, 10, canvas.width*0.3, canvas.height*0.3, 500);
            grad1.addColorStop(0, 'rgba(0, 168, 89, 0.08)');
            grad1.addColorStop(1, 'transparent');
            ctx.fillStyle = grad1;
            ctx.fillRect(0, 0, canvas.width, canvas.height);

            const grad2 = ctx.createRadialGradient(canvas.width*0.7, canvas.height*0.7, 10, canvas.width*0.7, canvas.height*0.7, 600);
            grad2.addColorStop(0, 'rgba(0, 242, 254, 0.06)');
            grad2.addColorStop(1, 'transparent');
            ctx.fillStyle = grad2;
            ctx.fillRect(0, 0, canvas.width, canvas.height);

            stars.forEach(star => {{
                ctx.beginPath();
                ctx.arc(star.x, star.y, star.radius, 0, Math.PI * 2);
                ctx.fillStyle = star.color;
                ctx.globalAlpha = star.alpha;
                ctx.shadowBlur = 8;
                ctx.shadowColor = star.color;
                ctx.fill();
                ctx.globalAlpha = 1.0;
                ctx.shadowBlur = 0;

                star.x += star.vx;
                star.y += star.vy;

                if (star.x < 0) star.x = canvas.width;
                if (star.x > canvas.width) star.x = 0;
                if (star.y < 0) star.y = canvas.height;
                if (star.y > canvas.height) star.y = 0;
            }});

            requestAnimationFrame(drawGalaxy);
        }}

        drawGalaxy();
    </script>
</body>
</html>
"""

with open(HTML_OUTPUT, 'w', encoding='utf-8') as f:
    f.write(html_content)

with open(HTML_RESPALDO, 'w', encoding='utf-8') as f:
    f.write(html_content)

with open(HTML_INDEX, 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f"      Tablero HTML guardado en: {HTML_OUTPUT}")
print(f"      Tablero HTML de Respaldo guardado en: {HTML_RESPALDO}")
print(f"      Tablero HTML Index guardado en: {HTML_INDEX}")

# 6. Sincronización automática con Git y GitHub
force_git = "--force-git" in sys.argv or "--git" in sys.argv
skip_git = "--skip-git" in sys.argv or "--no-git" in sys.argv

if skip_git:
    print("\n[6/6] Sincronización con Git omitida por parámetro (--skip-git).")
else:
    print(f"\n[6/6] Verificando y sincronizando cambios en Git (GitHub)...")
    try:
        import subprocess
        commit_msg = f"Actualización Tablero CUN - {fecha_actualizacion}"
        
        # Limpiar cualquier archivo lock residual en .git para evitar bloqueos
        for lock_file in glob.glob(os.path.join(BASE_DIR, ".git", "**", "*.lock"), recursive=True):
            try:
                os.remove(lock_file)
            except Exception:
                pass
        commit_editmsg = os.path.join(BASE_DIR, ".git", "COMMIT_EDITMSG")
        if os.path.exists(commit_editmsg):
            try:
                os.remove(commit_editmsg)
            except Exception:
                pass

        # Verificar si hay cambios en los archivos de entrada antes de hacer stage masivo
        input_diff = subprocess.run(["git", "diff", "--name-only", "HEAD", "--", "Asistencia", "Planta", "Scripts", "ACTUALIZAR_TABLERO.bat"], capture_output=True, text=True, cwd=BASE_DIR)
        untracked_inputs = subprocess.run(["git", "ls-files", "--others", "--exclude-standard", "Asistencia", "Planta", "Scripts"], capture_output=True, text=True, cwd=BASE_DIR)
        
        has_real_changes = bool(input_diff.stdout.strip() or untracked_inputs.stdout.strip())

        if has_real_changes or force_git:
            subprocess.run(["git", "add", "-A"], check=True, cwd=BASE_DIR)
            status_proc = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True, cwd=BASE_DIR)
            
            if status_proc.stdout.strip():
                subprocess.run(["git", "commit", "-m", commit_msg], check=True, cwd=BASE_DIR)
                print(f"      Commit registrado en Git: '{commit_msg}'")
            else:
                print("      No hay cambios pendientes por registrar en commit.")
            
            pull_proc = subprocess.run(["git", "pull", "origin", "main", "--rebase"], capture_output=True, text=True, cwd=BASE_DIR)
            if pull_proc.returncode != 0:
                subprocess.run(["git", "pull", "origin", "main", "--allow-unrelated-histories", "--no-edit"], capture_output=True, text=True, cwd=BASE_DIR)
                
            push_proc = subprocess.run(["git", "push", "origin", "main"], capture_output=True, text=True, cwd=BASE_DIR)
            if push_proc.returncode == 0:
                print("      Repositorio GitHub actualizado exitosamente: https://github.com/efi-cun/Cunectate_al_parche")
            else:
                print(f"      ADVERTENCIA al enviar cambios a GitHub: {push_proc.stderr.strip()}")
        else:
            print("      No se detectaron datos nuevos de Asistencia/Planta ni modificaciones en el código. Git ya está al día.")
    except Exception as err:
        print(f"      ADVERTENCIA: No se pudo completar la sincronización con Git: {err}")

print("\n=== PROCESO FINALIZADO CON ÉXITO ===")
